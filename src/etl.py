#######################normal code ##############################


import os
import re
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook

# Month mapping for textual dates
MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}

def is_time_related(value):
    if pd.isna(value):
        return False
    if isinstance(value, datetime):
        return True
    if isinstance(value, str):
        value = re.sub(r'\([^)]*\)', '', value).strip()
        date_patterns = [
            r'^\d{4}$',
            r'^\d{1,2}/\d{4}$',
            r'^\d{4}/\d{1,2}$',
            r'^\d{1,2}/\d{1,2}/\d{4}$',
            r'^[a-zA-Z]{3,}-\d{2}$',
        ]
        for pattern in date_patterns:
            if re.match(pattern, value):
                return True
        value_lower = value.lower()
        if any(month in value_lower for month in MONTHS):
            return True
    return False

def normalize_date(value, inferred_year=None):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = re.sub(r'\([^)]*\)', '', value).strip()
        try:
            if "-" in value or value.lower() in MONTHS:
                parts = value.split("-")
                month = MONTHS.get(parts[0].lower(), None)
                if month:
                    if len(parts) == 2 and parts[1].isdigit():
                        inferred_year = int("20" + parts[1]) if len(parts[1]) == 2 else int(parts[1])
                        return datetime(inferred_year, month, 1)
                    elif inferred_year:
                        return datetime(inferred_year, month, 1)
            if value.lower() in MONTHS and inferred_year:
                month = MONTHS[value.lower()]
                return datetime(inferred_year, month, 1)
        except Exception:
            pass
    return value

def process_time_structure(cells, inferred_year=None):
    dates = []
    current_year = inferred_year
    for cell in cells:
        if pd.isna(cell):
            break
        if is_time_related(cell):
            normalized_date = normalize_date(cell, current_year)
            if isinstance(normalized_date, datetime):
                current_year = normalized_date.year
            dates.append(normalized_date)
        elif current_year:
            normalized_date = normalize_date(cell, current_year)
            dates.append(normalized_date)
        else:
            dates.append(cell)
    return dates

def extract_time_data(sheet):
    time_data = []
    for row in sheet.iter_rows(values_only=True):
        inferred_year = None
        row_dates = process_time_structure(row, inferred_year)
        time_data.extend(row_dates)
    for col in sheet.iter_cols(values_only=True):
        inferred_year = None
        col_dates = process_time_structure(col, inferred_year)
        time_data.extend(col_dates)
    time_data = [d for d in time_data if isinstance(d, (datetime, str))]
    datetime_values = sorted([d for d in time_data if isinstance(d, datetime)])
    string_values = sorted([d for d in time_data if isinstance(d, str)])
    return datetime_values + string_values

def extract_parameters(sheet):
    parameters = {}
    for row in sheet.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if isinstance(cell, str) and not is_time_related(cell):
                values = row[i+1:]
                if all(isinstance(v, (int, float)) for v in values if v is not None):
                    parameters[cell] = values
    for col in sheet.iter_cols(values_only=True):
        for i, cell in enumerate(col):
            if isinstance(cell, str) and not is_time_related(cell):
                values = col[i+1:]
                if all(isinstance(v, (int, float)) for v in values if v is not None):
                    parameters[cell] = values
    return parameters

def save_organized_data(time_data, parameters, output_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Organized Data"
    sheet.cell(row=1, column=1, value="Date")
    for idx, time_value in enumerate(time_data, start=2):
        if isinstance(time_value, datetime):
            sheet.cell(row=idx, column=1, value=time_value.strftime('%Y-%m-%d'))
        else:
            sheet.cell(row=idx, column=1, value=time_value)
    col_idx = 2
    for param_name, values in parameters.items():
        sheet.cell(row=1, column=col_idx, value=param_name)
        for row_idx, value in enumerate(values, start=2):
            sheet.cell(row=row_idx, column=col_idx, value=value)
        col_idx += 1
    workbook.save(output_file)
    print(f"Saved: {output_file}")

def process_excel_file(input_path, output_folder):
    workbook = load_workbook(input_path, data_only=True)
    file_name = os.path.splitext(os.path.basename(input_path))[0]
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Processing: {file_name} - {sheet_name}")
        time_data = extract_time_data(sheet)
        parameters = extract_parameters(sheet)
        if time_data or parameters:
            output_file = os.path.join(output_folder, f"{file_name}_{sheet_name}_organized.xlsx")
            save_organized_data(time_data, parameters, output_file)
        else:
            print(f"No data found in: {file_name} - {sheet_name}")

def process_raw_excel_files(input_folder, intermediate_folder):
    os.makedirs(intermediate_folder, exist_ok=True)
    files = [f for f in os.listdir(input_folder) if f.lower().endswith('.xlsx')]
    if not files:
        print("No Excel files found in the input folder.")
    for file_name in files:
        input_path = os.path.join(input_folder, file_name)
        process_excel_file(input_path, intermediate_folder)

def process_excel_to_timeseries(input_directory, output_directory):
    os.makedirs(output_directory, exist_ok=True)
    consolidated_data = []
    for file_name in os.listdir(input_directory):
        if file_name.endswith(('.xlsx', '.xls')):
            file_path = os.path.join(input_directory, file_name)
            print(f"Processing file: {file_name}")
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                print(f"  Processing sheet: {sheet_name}")
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if 'Date' not in df.columns:
                    print(f"  Skipping sheet '{sheet_name}' as it does not contain a 'Date' column.")
                    continue
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                df = df.dropna(subset=['Date'])
                df = df.sort_values(by='Date')
                time_series_data = {}
                for col in df.columns:
                    if col != 'Date':
                        time_series_data[col] = [
                            {"Date": row['Date'].isoformat(), "value": row[col]}
                            for _, row in df[['Date', col]].dropna().iterrows()
                        ]
                json_structure = {
                    "metadata": {
                        "file_name": file_name,
                        "sheet_name": sheet_name
                    },
                    "data": time_series_data
                }
                json_output_path = os.path.join(output_directory, f"{file_name}_{sheet_name}.json")
                with open(json_output_path, 'w', encoding='utf-8') as json_file:
                    json.dump(json_structure, json_file, indent=4)
                print(f"    Saved JSON file: {json_output_path}")
                df.rename(columns={"Date": "Date"}, inplace=True)
                consolidated_data.append(df)
    if consolidated_data:
        combined_df = pd.concat(consolidated_data, ignore_index=True)
        excel_output_path = os.path.join(output_directory, "consolidated_timeseries.xlsx")
        combined_df.to_excel(excel_output_path, index=False)
        print(f"Saved consolidated Excel file: {excel_output_path}")
    else:
        print("No valid data found to save.")

def run_etl_pipeline(input_folder, intermediate_folder, output_folder):
    print("Step 1: Processing raw Excel files...")
    process_raw_excel_files(input_folder, intermediate_folder)
    print("Step 2: Converting organized Excel files to JSON and consolidated Excel...")
    process_excel_to_timeseries(intermediate_folder, output_folder)


################# multi-modal code ##############################

